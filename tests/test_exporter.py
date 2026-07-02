import csv
import io
from decimal import Decimal

from openpyxl import load_workbook

from app.utils.exporter import generate_csv, generate_excel
from src.models.monthly_tax import MonthlyTax


def _make_taxes() -> list[MonthlyTax]:
    return [
        MonthlyTax(
            periodo="2024-01",
            total_ventas=Decimal("1000000"),
            compras=Decimal("600000"),
            debito_fiscal=Decimal("190000"),
            credito_fiscal=Decimal("114000"),
            ppm=Decimal("50000"),
            ventas_afectas=Decimal("950000"),
            observaciones=[],
        ),
        MonthlyTax(
            periodo="2024-02",
            total_ventas=Decimal("2000000"),
            compras=Decimal("1200000"),
            debito_fiscal=Decimal("380000"),
            credito_fiscal=Decimal("228000"),
            ppm=Decimal("100000"),
            ventas_afectas=Decimal("1900000"),
            observaciones=["ajuste"],
        ),
        MonthlyTax(
            periodo="2024-03",
            total_ventas=Decimal("1500000"),
            compras=Decimal("800000"),
            debito_fiscal=Decimal("285000"),
            credito_fiscal=Decimal("152000"),
            ppm=Decimal("75000"),
            ventas_afectas=Decimal("1400000"),
            observaciones=["rectifica", "sin movimiento"],
        ),
    ]


class TestGenerateExcel:
    def test_generates_valid_workbook(self) -> None:
        taxes = _make_taxes()
        data = generate_excel(taxes, "2024-01", "2024-03")
        assert data

        wb = load_workbook(io.BytesIO(data))
        assert "Ventas_Compras" in wb.sheetnames
        ws = wb["Ventas_Compras"]

        headers = [cell.value for cell in ws[1]]
        assert "Periodo" in headers
        assert "Ventas" in headers
        assert "Compras" in headers
        assert "IVA Debito" in headers
        assert "IVA Credito" in headers
        assert "PPM" in headers
        assert "Ventas Netas" in headers
        assert "Observaciones" in headers

        assert ws["A2"].value == "2024-01"
        assert ws["B2"].value == "1.000.000"

        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None

    def test_headers_are_bold(self) -> None:
        taxes = _make_taxes()
        data = generate_excel(taxes, "2024-01", "2024-03")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Ventas_Compras"]

        for cell in ws[1]:
            assert cell.font.bold is True

    def test_filters_by_period(self) -> None:
        taxes = _make_taxes()
        data = generate_excel(taxes, "2024-02", "2024-02")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Ventas_Compras"]

        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][0] == "2024-02"

    def test_empty_range_returns_empty(self) -> None:
        taxes = _make_taxes()
        data = generate_excel(taxes, "2025-01", "2025-12")
        assert data == b""

    def test_observaciones_included(self) -> None:
        taxes = _make_taxes()
        data = generate_excel(taxes, "2024-01", "2024-03")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Ventas_Compras"]

        assert ws["H2"].value is None or ws["H2"].value == ""
        assert ws["H3"].value == "ajuste"
        assert ws["H4"].value == "rectifica; sin movimiento"


class TestGenerateCSV:
    def test_generates_valid_csv(self) -> None:
        taxes = _make_taxes()
        data = generate_csv(taxes, "2024-01", "2024-03")
        text = data.decode("utf-8")

        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert len(rows) == 4
        assert rows[0] == [
            "Periodo",
            "Ventas",
            "Compras",
            "IVA Debito",
            "IVA Credito",
            "PPM",
            "Ventas Netas",
            "Observaciones",
        ]
        assert rows[1][0] == "2024-01"
        assert rows[1][1] == "1.000.000"

    def test_filters_by_period(self) -> None:
        taxes = _make_taxes()
        data = generate_csv(taxes, "2024-03", "2024-03")
        text = data.decode("utf-8")
        rows = list(csv.reader(io.StringIO(text)))
        assert len(rows) == 2
        assert rows[1][0] == "2024-03"

    def test_empty_range_returns_empty(self) -> None:
        taxes = _make_taxes()
        data = generate_csv(taxes, "2025-01", "2025-12")
        assert data == b""

    def test_number_format_chilean(self) -> None:
        taxes = _make_taxes()
        data = generate_csv(taxes, "2024-01", "2024-01")
        text = data.decode("utf-8")
        rows = list(csv.reader(io.StringIO(text)))
        assert rows[1][1] == "1.000.000"

    def test_no_scientific_notation(self) -> None:
        taxes = [
            MonthlyTax(
                periodo="2024-01",
                total_ventas=Decimal("1000000000"),
                compras=Decimal("500000000"),
                debito_fiscal=Decimal("190000000"),
                credito_fiscal=Decimal("95000000"),
                ppm=Decimal("50000000"),
                ventas_afectas=Decimal("950000000"),
                observaciones=[],
            )
        ]
        data = generate_csv(taxes, "2024-01", "2024-01")
        text = data.decode("utf-8")
        assert "1.000.000.000" in text
        assert "e+" not in text.lower()

    def test_utf8_encoding(self) -> None:
        taxes = [
            MonthlyTax(
                periodo="2024-01",
                total_ventas=Decimal("1000"),
                compras=Decimal("500"),
                debito_fiscal=Decimal("190"),
                credito_fiscal=Decimal("95"),
                ppm=Decimal("50"),
                ventas_afectas=Decimal("950"),
                observaciones=["débito fiscal"],
            )
        ]
        data = generate_csv(taxes, "2024-01", "2024-01")
        text = data.decode("utf-8")
        assert "débito" in text
