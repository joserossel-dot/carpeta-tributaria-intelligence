import csv
import io
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.models.monthly_tax import MonthlyTax


def _fmt(val: Decimal | None) -> str:
    if val is None:
        return ""
    return f"{val:,.0f}".replace(",", ".")


def _fmt_obs(obs: list[str]) -> str:
    return "; ".join(obs) if obs else ""


def _build_rows(
    monthly_taxes: list[MonthlyTax],
    desde: str,
    hasta: str,
) -> list[dict[str, str]]:
    headers = [
        "Periodo",
        "Ventas",
        "Compras",
        "IVA Debito",
        "IVA Credito",
        "PPM",
        "Ventas Netas",
        "Observaciones",
    ]
    rows: list[dict[str, str]] = []
    for mt in monthly_taxes:
        if desde <= mt.periodo <= hasta:
            rows.append({
                "Periodo": mt.periodo,
                "Ventas": _fmt(mt.total_ventas),
                "Compras": _fmt(mt.compras),
                "IVA Debito": _fmt(mt.debito_fiscal),
                "IVA Credito": _fmt(mt.credito_fiscal),
                "PPM": _fmt(mt.ppm),
                "Ventas Netas": _fmt(mt.ventas_afectas),
                "Observaciones": _fmt_obs(mt.observaciones),
            })
    return [headers] + rows


def generate_excel(
    monthly_taxes: list[MonthlyTax],
    desde: str,
    hasta: str,
) -> bytes:
    rows = _build_rows(monthly_taxes, desde, hasta)
    if len(rows) < 2:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas_Compras"

    bold = Font(bold=True)

    for col_idx, header in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, key in enumerate(rows[0], start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])

    ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, len(rows[0]) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_csv(
    monthly_taxes: list[MonthlyTax],
    desde: str,
    hasta: str,
) -> bytes:
    rows = _build_rows(monthly_taxes, desde, hasta)
    if len(rows) < 2:
        return b""

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(rows[0])
    for row in rows[1:]:
        writer.writerow([row[key] for key in rows[0]])

    return buf.getvalue().encode("utf-8")
